"""
    Author: Pickolzi
    Purpose: Scrape data from the anime website, myanimelist.net, through a CustomGUI
"""
import requests
from openpyxl import Workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from bs4 import BeautifulSoup
from anime import Anime

LINK = "https://myanimelist.net/topanime.php"

def retrieve_data(link="https://myanimelist.net/topanime.php"):
    """
    :param link: grabs html from the link.
    :return: list of Anime objects with their attributes.
    """
    html = requests.get(link).text
    soup = BeautifulSoup(html, "html.parser")
    table = soup.table

    animes = []

    # Collecting individual anime attributes.
    ranking_list = table.find_all(class_="ranking-list")
    for anime in ranking_list:
        info = anime.find(class_="information di-ib mt4").text
        info = info.split("\n")
        info = [i.strip() for i in info if i.strip() != ""]

        rank = int(anime.find(class_="rank ac").text.strip())
        title = anime.h3.text
        anime_type = info[0].split(" ")[0]
        num_of_episodes = info[0].split("(")[1].split(" ")[0]
        release_date = info[1]
        members = info[2].split(" ")[0]
        try:
            score = float(anime.find(class_="score ac fs14").span.text)
        except ValueError:
            score = "N/A"
        link = anime.h3.a["href"]


        anime_object = Anime(rank, title, anime_type, num_of_episodes, release_date, members, score, link)
        animes.append(anime_object)

    return animes


def grab_image_from_google(url):
    """
    :param title: url for MAL anime page.
    :return: url for #1 ranked anime image
    """
    html = requests.get(url).text
    soup = BeautifulSoup(html, "html.parser")
    query = soup.table.find_all("tr")[1].h3.a["href"]

    html = requests.get(query).text
    soup = BeautifulSoup(html, "html.parser")
    image_url = soup.table.img["data-src"]

    return image_url


def save_to_excel(animes, sort_method):
    """
    :param animes: Anime object
    Creates xlsx file and saves the Anime object's data to the spreadsheet.
    """
    filename = "results.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Top Animes"

    animes = sort_animes(animes, sort_method)

    # Create header information
    ws["A1"] = "Rank:"
    ws["B1"] = "Title:"
    ws["C1"] = "Anime type:"
    ws["D1"] = "Number of episodes:"
    ws["E1"] = "Release date:"
    ws["F1"] = "Number of members:"
    ws["G1"] = "Score:"
    ws["H1"] = "Link:"

    # Add anime information
    for index, anime in enumerate(animes):
        ws[f"A{index+2}"] = anime.rank
        ws[f"B{index+2}"] = anime.title
        ws[f"C{index+2}"] = anime.anime_type
        ws[f"D{index+2}"] = anime.num_of_episodes
        ws[f"E{index+2}"] = anime.release_date
        ws[f"F{index+2}"] = anime.members
        ws[f"G{index+2}"] = anime.score
        ws[f"H{index+2}"] = anime.link

    # Configuration settings
    ws.column_dimensions["B"].width = 64
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 64

    # Close the workbook
    wb.save(filename=filename)
    print("Finished saving to xlsx file...")


def sort_animes(animes, sort_method):
    altered_anime_list = []

    if sort_method == "rank":
        altered_anime_list = animes

    elif sort_method == "title":
        titles = {}
        for index, anime in enumerate(animes):
            key = anime.title

            # A bit of cleaning
            if key[0] == '\"':
                key = key[1:-1]
            titles[key] = index

        # Creates a list of sorted anime titles
        titles_list_order = sorted(titles)
        for title in titles_list_order:
            # Grabs anime by alphabetical order.
            altered_anime_list.append(animes[titles[title]])

    elif sort_method == "anime type":
        anime_types = ["TV", "Movie", "OVA", "ONA"]
        temp_lists = [list() for _ in anime_types]
        for anime in animes:
            try:
                index = anime_types.index(anime.anime_type)
                temp_lists[index].append(anime)
            except ValueError:
                continue

        # Combining the 4 lists into 1: altered_anime_list
        for temp_list in temp_lists:
            altered_anime_list.extend(temp_list)

    elif sort_method == "episodes":
        temp_animes_list = []
        episodes_list = []

        end_animes = []

        for anime in animes:
            if anime.num_of_episodes == "?":
                end_animes.append(anime)
                continue

            current_episode = int(anime.num_of_episodes)
            for episode in episodes_list:
                if current_episode <= episode:
                    index = episodes_list.index(episode)
                    temp_animes_list.insert(index, anime)
                    episodes_list.insert(index, current_episode)
                    break
            else:
                temp_animes_list.append(anime)
                episodes_list.append(current_episode)

        altered_anime_list = temp_animes_list + end_animes

    elif sort_method == "release date":
        pass

    elif sort_method == "members":
        pass

    elif sort_method == "score":
        pass

    return altered_anime_list


if __name__ == '__main__':

    animes = retrieve_data("https://myanimelist.net/topanime.php")
    # animes = retrieve_data("https://myanimelist.net/topanime.php?type=upcoming")
    # save_to_excel(animes)


    sorting_options = """
    rank
    title
    anime type
    episodes
    release date
    members
    score
    """
    temp_options = sorting_options.strip().split("\n")
    # animes = sort_animes(animes, temp_options[1])
    animes = sort_animes(animes, "title")
    for anime in animes:
        print(anime.title)

