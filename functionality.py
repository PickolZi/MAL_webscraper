"""
    Author: Pickolzi
    Purpose: Scrape data from the anime website, myanimelist.net, through a CustomGUI
"""
import requests
from openpyxl import Workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from bs4 import BeautifulSoup
from anime import Anime

LINK = "https://myanimelist.net/topanime.php"

def retrieve_data(link="https://myanimelist.net/topanime.php"):
    """
    :param link: grabs html from the link.
    :return: list of Anime objects with their attributes.
    """
    html = requests.get(link).text
    soup = BeautifulSoup(html, "html.parser")
    table = soup.table

    animes = []

    # Collecting individual anime attributes.
    ranking_list = table.find_all(class_="ranking-list")
    for anime in ranking_list:
        info = anime.find(class_="information di-ib mt4").text
        info = info.split("\n")
        info = [i.strip() for i in info if i.strip() != ""]

        rank = int(anime.find(class_="rank ac").text.strip())
        title = anime.h3.text
        anime_type = info[0].split(" ")[0]
        num_of_episodes = info[0].split("(")[1].split(" ")[0]
        release_date = info[1]
        members = info[2].split(" ")[0]
        try:
            score = float(anime.find(class_="score ac fs14").span.text)
        except ValueError:
            score = "N/A"
        link = anime.h3.a["href"]


        anime_object = Anime(rank, title, anime_type, num_of_episodes, release_date, members, score, link)
        animes.append(anime_object)

    return animes


def grab_image_from_google(url):
    """
    :param title: url for MAL anime page.
    :return: url for #1 ranked anime image
    """
    html = requests.get(url).text
    soup = BeautifulSoup(html, "html.parser")
    query = soup.table.find_all("tr")[1].h3.a["href"]

    html = requests.get(query).text
    soup = BeautifulSoup(html, "html.parser")
    image_url = soup.table.img["data-src"]

    return image_url


def save_to_excel(animes, sort_method, filename):
    """
    :param animes: Anime object
    Creates xlsx file and saves the Anime object's data to the spreadsheet.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "Top Animes"

    animes = sort_animes(animes, sort_method)

    # Create header information
    ws["A1"] = "Rank:"
    ws["B1"] = "Title:"
    ws["C1"] = "Anime type:"
    ws["D1"] = "Number of episodes:"
    ws["E1"] = "Release date:"
    ws["F1"] = "Number of members:"
    ws["G1"] = "Score:"
    ws["H1"] = "Link:"

    # Add anime information
    for index, anime in enumerate(animes):
        ws[f"A{index+2}"] = anime.rank
        ws[f"B{index+2}"] = anime.title
        ws[f"C{index+2}"] = anime.anime_type
        ws[f"D{index+2}"] = anime.num_of_episodes
        ws[f"E{index+2}"] = anime.release_date
        ws[f"F{index+2}"] = anime.members
        ws[f"G{index+2}"] = anime.score
        ws[f"H{index+2}"] = anime.link

    # Configuration settings
    ws.column_dimensions["B"].width = 64
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 64

    # Close the workbook
    wb.save(filename=filename)
    print("Finished saving to xlsx file...")


def sort_animes(animes, sort_method):
    altered_anime_list = []

    if sort_method == "rank":
        altered_anime_list = animes

    elif sort_method == "title":
        titles = {}
        for index, anime in enumerate(animes):
            key = anime.title

            # A bit of cleaning
            if key[0] == '\"':
                key = key[1:-1]
            titles[key] = index

        # Creates a list of sorted anime titles
        titles_list_order = sorted(titles)
        for title in titles_list_order:
            # Grabs anime by alphabetical order.
            altered_anime_list.append(animes[titles[title]])

    elif sort_method == "anime type":
        anime_types = ["TV", "Movie", "OVA", "ONA"]
        temp_lists = [list() for _ in anime_types]
        for anime in animes:
            try:
                index = anime_types.index(anime.anime_type)
                temp_lists[index].append(anime)
            except ValueError:
                continue

        # Combining the 4 lists into 1: altered_anime_list
        for temp_list in temp_lists:
            altered_anime_list.extend(temp_list)

    elif sort_method == "episodes":
        temp_animes_list = []
        episodes_list = []

        end_animes = []

        for anime in animes:
            if anime.num_of_episodes == "?":
                end_animes.append(anime)
                continue

            current_episode = int(anime.num_of_episodes)
            for episode in episodes_list:
                if current_episode <= episode:
                    index = episodes_list.index(episode)
                    temp_animes_list.insert(index, anime)
                    episodes_list.insert(index, current_episode)
                    break
            else:
                temp_animes_list.append(anime)
                episodes_list.append(current_episode)

        altered_anime_list = temp_animes_list + end_animes

    elif sort_method == "release date":
        temp_anime_list = []
        temp_release_date = []
        end_animes = []

        MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        for anime in animes:
            current_release_date = anime.release_date.split(" ")[:2]

            # No release date, add to end of the list.
            if current_release_date[0] == "-":
                end_animes.append(anime)
                continue

            # Month and year of the anime release date.
            year = current_release_date[1]
            month = current_release_date[0]

            # Release dates with no months.
            if month not in MONTHS:
                year = int(month)  # This is year, just the format is different when scraping.
                month = "N/A"
                end_animes.insert(0, anime)  # If there's no month, that means it's made in this recent year, so it should be put towards the beginning of the end_animes list.
                continue
            if month != "N/A":
                year = int(year)
                month = MONTHS.index(month)

            # Sort the dates by month that do have both the months and years
            for date in temp_release_date:
                if month <= date[0]:
                    index = temp_release_date.index(date)
                    temp_release_date.insert(index, (month, year))
                    temp_anime_list.insert(index, anime)
                    break
            else:
                temp_release_date.append((month, year))
                temp_anime_list.append(anime)

        final_anime_list = []
        final_release_date = []

        # Sorting by year
        for date, anime in zip(temp_release_date, temp_anime_list):
            current_year = date[1]
            for tuple in final_release_date:
                year = tuple[1]
                if current_year <= year:
                    index = [x[1] for x in final_release_date].index(year)
                    if current_year == year:
                        skip = [x[1] for x in final_release_date].count(year)
                        index += skip
                    final_anime_list.insert(index, anime)
                    final_release_date.insert(index, date)
                    break
            else:
                final_anime_list.append(anime)
                final_release_date.append(date)

        altered_anime_list = final_anime_list + end_animes

    elif sort_method == "members":
        temp_anime_list = []
        temp_member_count_list = []

        for anime in animes:
            # Turns current member from string to int.
            current_member_count = anime.members.split(",")
            current_member_count = "".join(current_member_count)
            current_member_count = int(current_member_count)

            # Sorts the anime objects and member counts into the temporary lists.
            for member_count in temp_member_count_list:
                if current_member_count <= member_count:
                    index = temp_member_count_list.index(member_count)
                    temp_anime_list.insert(index, anime)
                    temp_member_count_list.insert(index, current_member_count)
                    break
            else:
                temp_anime_list.append(anime)
                temp_member_count_list.append(current_member_count)

        altered_anime_list = temp_anime_list

    elif sort_method == "score":
        temp_anime_list = []
        temp_score_list = []

        for anime in animes:
            current_score = anime.score
            for score in temp_score_list:
                if current_score <= score:
                    index = temp_score_list.index(score)
                    temp_anime_list.insert(index, anime)
                    temp_score_list.insert(index, current_score)
                    break
            else:
                temp_anime_list.append(anime)
                temp_score_list.append(current_score)

        altered_anime_list = temp_anime_list

    return altered_anime_list
